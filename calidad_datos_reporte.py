import os
import glob
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import dask.dataframe as dd
from dask.diagnostics import ProgressBar
from pathlib import Path
from openpyxl import load_workbook

# ==========================================
# 1. CONFIGURACIÓN Y PARÁMETROS
# ==========================================
RANGOS_VALIDACION = {
    # ————————————— Identificadores / Trazabilidad —————————————
    'RPM': (500, 2000),
    'Actual Speed (RPM)': (500, 2000),
    'Turbocharger Speed (RPM)': (0, 70000),
    'Velocidad (Km/h)': (0, 60),
    'velocidad_km_h': (0, 60),
    
    # ————————————— Torque / Carga / Potencia —————————————
    'Actual Percent Torque (%)': (0, 100),
    'Percent Load At Current Speed (%)': (0, 100),
    'Engine Demand - Percent Torque (%)': (0, 100),
    'F. de Carga': (0, 100),
    'Pedal': (0, 100),
    'Accelerator position (%)': (0, 100),
    'Power (HP)': (0, 3000),
    'Potencia': (0, 3000),
    'potencia': (0, 3000),
    "Engine Torque Mode ()": (0, 100),

    # ————————————— Combustible —————————————
    'Fuel rate (L/h)': (0, 800),
    'Engine Average Fuel Economy (Kilometro por Litro)': (0.1, 3),
    'Engine Instantaneous Fuel Economy (Kilometro por Litro)': (0.1, 3),
    'Fuel Temperature (F)': (30, 195),
    'Engine Fuel Delivery Pressure (Extended Range) (PSI)': (50, 160),
    'Fuel Delivery Pressure (PSI)': (50, 160),
    'Fuel Supply Pump Inlet Pressure (MCRS) (PSI)': (10, 120),
    'Engine Fuel Supply Pump Intake Pressure (Extended Range) (PSI)': (10, 120),
    'Injector Metering (PSI)': (1000, 30000),

    # ————————————— Presiones de admisión / turbo —————————————
    'IMP (PSI)': (14.5, 50),
    'IMP-LB (PSI)': (14.5, 50),
    'IMP-RB (MCRS) (PSI)': (14.5, 50),
    'Barometric Pressure (PSI)': (11.6, 15.9),
    'Crankcase Pressure (MCRS) (in-H2O)': (-15, 15),
    'Coolant Pressure (PSI)': (5, 25),
    'Oil Differential Pressure (PSI)': (10, 80),
    'Pre-filter Oil Pressure (PSI)': (20, 90),
    'Post Oil Filter (MCRS) (PSI)': (20, 90),
    'Rifle Oil Pressure (PSI)': (20, 100),
    'Engine fuel Pump Intake Oil Pressure (kPa)': (100, 700),

    # ————————————— Temperaturas —————————————
    'Coolant temperature (F)': (140, 240),
    'Engine Oil Temperature (F)': (120, 250),
    'Oil temperature (FPS 2) (F)': (120, 250),
    'Oil Temp (FPS 2) (F)': (120, 250),
    'Ecu temperature (F)': (70, 200),
    'IMT-RBR (F)': (50, 250),
    'IMT-RBF (F)': (50, 250),
    'IMT-LBR (F)': (50, 250),
    'IMT-LBF (F)': (50, 250),
    'EGT-AV (F)': (400, 1300),
    'EGT-LB (MCRS) (F)': (400, 1300),
    'EGT-RB (MCRS) (F)': (400, 1300),
    **{f'EGT-{i:02d} (F)': (400, 1300) for i in range(1, 17)}, # EGT-01 a EGT-16

    # ————————————— Niveles y estados —————————————
    'Engine coolant level (%)': (0, 100),
    'Oil Level - Reserve Tank ()': (0, 100),

    # ————————————— Eléctricos —————————————
    'Electrical potential (V)': (20, 32),
    'Battery potential (V)': (20, 32),
    'Ecu voltage (V)': (20, 32),

    # ————————————— GPS / Ubicación —————————————
    'Latitud': (-34, -32),
    'Longitud': (-71, -69.0),
    'latitud': (-34, -32),
    'longitud': (-71, -69.0),
    'Altitude (m)': (2700, 3800),
}

def analizar_calidad_con_dask(file_path, rangos_validacion):
    print(f"  -> Analizando con Dask: {os.path.basename(file_path)}")

    # 1. Leer el archivo con Dask.
    # Dask lee el archivo de forma "perezosa" (lazy), sin cargarlo en memoria.
    # 'blocksize' es similar a 'chunksize'. 25MB es un buen punto de partida.
    # 'dtype=str' sigue siendo la mejor práctica para evitar errores de tipo.
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext == '.csv':
            ddf = dd.read_csv(file_path, sep=None, engine='python', dtype=str, blocksize="25MB", 
            assume_missing=True, on_bad_lines='warn')
        elif file_ext in ['.xlsx', '.xls', '.xlsm']:
            # Para Excel, usamos una estrategia de streaming para no agotar la memoria.
            # Leemos el archivo en trozos con pandas y los convertimos a Dask.
            def stream_excel_chunks(path, chunk_size=20000):
                wb = load_workbook(filename=path, read_only=True, data_only=True)
                ws = wb.active
                rows = ws.iter_rows()
                header = [cell.value for cell in next(rows)]
                chunk = []
                for row in rows:
                    chunk.append([cell.value for cell in row])
                    if len(chunk) >= chunk_size:
                        yield pd.DataFrame(chunk, columns=header)
                        chunk = []
                if chunk:
                    yield pd.DataFrame(chunk, columns=header)

            # evitando cargar todo el archivo en memoria a la vez.
            bag = dd.from_sequence(stream_excel_chunks(file_path))
            ddf = bag.to_dataframe()
    except Exception as e:
        print(f" Error al leer el archivo con Dask: {e}")
        return pd.DataFrame()

    # 2. Definir los cálculos (aún no se ejecutan)
    total_filas = len(ddf)

    # Contar nulos por columna
    nulos_por_columna = ddf.isnull().sum()

    # Contar ceros y valores fuera de rango
    ceros_por_columna = {}
    fuera_rango_por_columna = {}
    
    for col in ddf.columns:
        col_numerica = dd.to_numeric(ddf[col], errors='coerce')
        
        
        # para evitar el AttributeError en el objeto 'ToNumeric' de Dask.
        ceros_por_columna[col] = (col_numerica.notnull() & (col_numerica == 0)).sum()
        
        # Si la columna tiene un rango de validación definido
        if col in rangos_validacion:
            lo, hi = rangos_validacion[col]
            # Un valor fuera de rango no debe ser ni nulo ni cero.
            no_nulos_ni_ceros = col_numerica.notnull() & (col_numerica != 0)
            fuera_rango_por_columna[col] = ((col_numerica[no_nulos_ni_ceros] < lo) | (col_numerica[no_nulos_ni_ceros] > hi)).sum()
        else:
            # Si no hay rango, no hay valores fuera de rango
            fuera_rango_por_columna[col] = ddf[col].map_partitions(lambda s: 0, meta=(col, int)).sum()

    # 3. Ejecutar todos los cálculos en paralelo con .compute()
    print("  -> Ejecutando cálculos en paralelo...")
    with ProgressBar():
        # dask.compute puede ejecutar múltiples cálculos a la vez
        total_filas_res, nulos_res, ceros_res, fuera_rango_res = dd.compute(
            total_filas, nulos_por_columna, ceros_por_columna, fuera_rango_por_columna)

    # Ahora que tenemos el resultado, comprobamos si el archivo estaba vacío.
    if total_filas_res == 0:
        print("  El archivo está vacío o no se pudo leer.")
        return pd.DataFrame()

    print("Cálculos completados. Generando informe...")

    # 4. Ensamblar el informe final (esto ya se hace en memoria con los resultados)
    resumen_calidad = []
    for columna in ddf.columns:
        nulos = nulos_res[columna]
        ceros = ceros_res[columna]
        fuera_rango = fuera_rango_res.get(columna, 0) # .get() para seguridad
        
        resumen_calidad.append({
            'Variable': columna,
            'Total_Filas': total_filas_res,
            'Nulos': nulos,
            'Ceros': ceros,
            'Fuera_Rango': fuera_rango,
        })

    df_resumen = pd.DataFrame(resumen_calidad)
    return df_resumen


# ==========================================
# 3. GENERACIÓN DE REPORTES Y GRÁFICOS
# ==========================================
def generar_graficos_globales(final_df, output_reports_dir, best_trucks_file_path):
    """
    Genera los gráficos consolidados y determina los mejores camiones.
    """
    print("\n\n---  Generando Gráficos Consolidados para TODA la Flota ---")
    
    # 2. Agrupar por variable y sumar los conteos a nivel global
    global_sum_report = final_df.groupby('Variable').sum().reset_index()
    
    # 3. Calcular los porcentajes finales para toda la flota
    total_filas_global = global_sum_report['Total_Filas']
    invalidos_global = global_sum_report['Nulos'] + global_sum_report['Ceros'] + global_sum_report['Fuera_Rango']
    validos_global = (total_filas_global - invalidos_global).clip(lower=0)
    global_sum_report['Validos_%'] = (validos_global / total_filas_global* 100).round(2)
    
    global_report = global_sum_report.sort_values('Validos_%', ascending=False)

    # Guardar el resumen global en CSV con porcentajes
    resumen_global_path = output_reports_dir / "resumen_global_calidad.csv"
    global_report.to_csv(resumen_global_path, index=False)
    print(f"\n Resumen global de porcentajes guardado en: {resumen_global_path}")

    # --- 4. Generar gráfico de CALIDAD GPS POR CAMIÓN ---
    variables_gps = ['Latitud', 'Longitud', 'Altitude (m)', 'latitud', 'longitud']
    
    # Calculamos el % de validez de 'Latitud' para cada camión
    reporte_gps_por_camion = final_df[final_df['Variable'].isin(['Latitud', 'latitud'])].copy()
    if not reporte_gps_por_camion.empty:
        total_filas_camion = reporte_gps_por_camion['Total_Filas']
        invalidos_camion = reporte_gps_por_camion['Nulos'] + reporte_gps_por_camion['Ceros'] + reporte_gps_por_camion['Fuera_Rango']
        validos_camion = (total_filas_camion - invalidos_camion).clip(lower=0)
        reporte_gps_por_camion['Validos_%'] = (validos_camion / total_filas_camion * 100).round(2)
        reporte_gps_por_camion = reporte_gps_por_camion.sort_values('Validos_%', ascending=False)

        plt.figure(figsize=(max(12, len(reporte_gps_por_camion) * 0.5), 7))
        plt.bar(reporte_gps_por_camion['camion'], reporte_gps_por_camion['Validos_%'], color='royalblue')
        plt.title('Calidad de Coordenadas GPS por Camión', fontsize=16)
        plt.ylabel('% de Coordenadas Válidas', fontsize=12)
        plt.xlabel('Camión', fontsize=12)
        plt.xticks(rotation=45, ha="right")
        plt.ylim(0, 100)
        plt.grid(axis='y', linestyle='--', alpha=0.7)
        plt.tight_layout()
        plt.savefig(output_reports_dir / 'calidad_consolidada_GPS_por_Camion.png', dpi=150, bbox_inches='tight')
        plt.close()
        print("Gráfico de calidad de GPS por camión guardado.")

        # 1. Encontrar el máximo porcentaje de validez
        max_calidad_gps = reporte_gps_por_camion['Validos_%'].max()
        
        # 1. Calcular un score de calidad que combine GPS y Fuel Rate
        reporte_seleccion = reporte_gps_por_camion[['camion', 'Validos_%']].rename(columns={'Validos_%': 'GPS_Validos_%'})

        # Extraemos la validez del Fuel Rate para cada camión
        reporte_fuel_rate = final_df[final_df['Variable'] == 'Fuel rate (L/h)'].copy()
        if not reporte_fuel_rate.empty:
            total_filas_fuel = reporte_fuel_rate['Total_Filas']
            invalidos_fuel = reporte_fuel_rate['Nulos'] + reporte_fuel_rate['Ceros'] + reporte_fuel_rate['Fuera_Rango']
            validos_fuel = (total_filas_fuel - invalidos_fuel).clip(lower=0)
            reporte_fuel_rate['Fuel_Validos_%'] = (validos_fuel / total_filas_fuel * 100).round(2)
            
            # Unimos ambos reportes
            reporte_seleccion = pd.merge(reporte_seleccion, reporte_fuel_rate[['camion', 'Fuel_Validos_%']], on='camion', how='left').fillna(0)
            
            # Calculamos el score (ej: producto de ambos porcentajes)
            reporte_seleccion['Score_Calidad'] = reporte_seleccion['GPS_Validos_%'] * reporte_seleccion['Fuel_Validos_%']
        else:
            reporte_seleccion['Score_Calidad'] = reporte_seleccion['GPS_Validos_%']

        # 2. Encontrar el máximo score y los camiones que lo tienen
        max_score = reporte_seleccion['Score_Calidad'].max()
        mejores_camiones_df = reporte_seleccion[reporte_seleccion['Score_Calidad'] == max_score]
        lista_mejores_camiones = mejores_camiones_df['camion'].tolist()

        # 3. Guardar los nombres en un archivo de texto
        best_trucks_file_path.parent.mkdir(parents=True, exist_ok=True)
        with open(best_trucks_file_path, 'w') as f:
            for nombre_camion in lista_mejores_camiones:
                f.write(f"{nombre_camion}\n")

        print(f"\n\n---  Mejor(es) Camión(es) Identificado(s) ---")
        print(f"Los camiones con el mejor score de calidad ({max_score:.2f}) son: {', '.join(lista_mejores_camiones)}")
        print(f" Sus nombres han sido guardados en: {best_trucks_file_path}")

    else:
        print("\n\n---No se pudo determinar un mejor camión (no se procesaron datos GPS). ---")

# ==========================================
# 4. ORQUESTADOR PRINCIPAL
# ==========================================
def run_quality_analysis(base_data_path, output_reports_path,
                          best_trucks_file_path, camiones_a_procesar_list=None):
    """
    Función principal que ejecuta el análisis de calidad de datos.
    """
    # --- CONFIGURACIÓN ---
    BASE_PATH = Path(base_data_path)
    OUTPUT_REPORTS_DIR = Path(output_reports_path)
    OUTPUT_REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    # Define qué carpetas (camiones) procesar. Si la lista está vacía, se procesan todas.
    if camiones_a_procesar_list is None:
        CAMIONES_A_PROCESAR = []
    else:
        CAMIONES_A_PROCESAR = camiones_a_procesar_list

    if not os.path.isdir(BASE_PATH):
        raise FileNotFoundError(f" Error: La ruta base '{BASE_PATH}' no existe.")

    # Si la lista está vacía, busca todas las subcarpetas en la ruta base
    if not CAMIONES_A_PROCESAR:
        CAMIONES_A_PROCESAR = [d for d in os.listdir(BASE_PATH) if 
                               os.path.isdir(os.path.join(BASE_PATH, d))]
        print(f" No se especificaron camiones: {len(CAMIONES_A_PROCESAR)}.")

    all_consolidated_reports = []

    for camion in CAMIONES_A_PROCESAR:
        camion_path = os.path.join(BASE_PATH, camion)
        if not os.path.isdir(camion_path):
            print(f"  Advertencia: La carpeta para el camión '{camion}' no se encontró. Saltando...")
            continue

        # Crear subcarpeta de salida para los reportes del camión
        output_camion_dir = os.path.join(OUTPUT_REPORTS_DIR, camion)
        os.makedirs(output_camion_dir, exist_ok=True)

        print(f"\n--- Procesando Camión: {camion} ---")
        
        quality_reports_por_camion = []
        # Buscar archivos CSV y Excel recursivamente
        search_patterns = [os.path.join(camion_path, '**', f'*.{ext}') 
                           for ext in ['csv', 'xlsx', 'xls']]
        files_to_process = []
        for pattern in search_patterns:
            files_to_process.extend(glob.glob(pattern, recursive=True))

        if not files_to_process:
            print("  -> No se encontraron archivos .csv o .xlsx para analizar.")
            continue

        for file_path in files_to_process:
            print(f"\nProcesando archivo: {os.path.basename(file_path)}")
            print("-" * 60)
            
            quality_df = analizar_calidad_con_dask(file_path, RANGOS_VALIDACION)
            
            if not quality_df.empty:
                print("\n--- Resumen de Calidad de Datos (Archivo Individual) ---")
                print(quality_df.to_string(index=False))
                quality_reports_por_camion.append(quality_df)
            else:
                print("  -> No se generó el informe de calidad (archivo vacío o error de lectura).")

        # --- Consolidación y Gráfico por Camión ---
        if quality_reports_por_camion:
            # Concatenar todos los reportes del camión y calcular el promedio por variable
            # 1. Concatenar todos los reportes (que ahora tienen conteos)
            consolidated_df = pd.concat(quality_reports_por_camion)
            
            # 2. Agrupar por variable y sumar los conteos absolutos
            sum_report = consolidated_df.groupby('Variable').sum().reset_index()
            
            # 3. Calcular los porcentajes finales a partir de las sumas totales
            total_filas_camion = sum_report['Total_Filas']
            sum_report['Nulos_%'] = (sum_report['Nulos'] / total_filas_camion * 100).round(2)
            sum_report['Ceros_%'] = (sum_report['Ceros'] / total_filas_camion * 100).round(2)
            sum_report['Fuera_Rango_%'] = (sum_report['Fuera_Rango'] / total_filas_camion * 100).round(2)
            
            invalidos = sum_report['Nulos'] + sum_report['Ceros'] + sum_report['Fuera_Rango']
            validos = (total_filas_camion - invalidos).clip(lower=0)
            sum_report['Validos_%'] = (validos / total_filas_camion * 100).round(2)
            
            consolidated_report = sum_report
            consolidated_report = consolidated_report.sort_values('Validos_%', ascending=False)
            
            print(f"\n\n---  Resumen Consolidado para Camión: {camion} ---")
            print(consolidated_report.to_string(index=False))

            # Guardar el resumen consolidado del camión en CSV con porcentajes
            resumen_camion_path = os.path.join(output_camion_dir, f"{camion}_resumen_calidad.csv")
            consolidated_report.to_csv(resumen_camion_path, index=False)
            print(f"  -> Resumen de porcentajes guardado en: {resumen_camion_path}")

            # Añadir el reporte de sumas (no de porcentajes) a la lista global
            sum_report['camion'] = camion # Añadimos el nombre del camión para el reporte final
            all_consolidated_reports.append(sum_report)

            # La generación de gráficos ha sido desactivada según la solicitud.
            print(f"\n Análisis consolidado para '{camion}' completado. No se generará gráfico.")

    # --- GRÁFICOS FINALES CONSOLIDADOS PARA TODA LA FLOTA ---
    if all_consolidated_reports:
        # 1. Unir todos los reportes de sumas de todos los camiones
        final_df = pd.concat(all_consolidated_reports, ignore_index=True)
        generar_graficos_globales(final_df, OUTPUT_REPORTS_DIR, best_trucks_file_path)

    print("\n\n Proceso de análisis de calidad completado.")


if __name__ == "__main__":
    # --- Rutas para ejecución directa ---
    project_base_path = Path(r"C:\Users\icquerov\OneDrive - Anglo American\Desktop\Proyecto_Caminos")
    base_camiones_path = project_base_path / "camiones"
    output_reports_dir = project_base_path / "reportes_calidad"
    
    # Archivo donde se guardará el resultado
    output_dir_main = project_base_path / "outputs"
    best_trucks_file = output_dir_main / "mejores_camiones.txt"

    # Para probar con un solo camión, pon su nombre aquí. Para procesar todos, deja la lista vacía [].
    CAMIONES_PARA_PRUEBA = [] # Ej: ["CDH76"] para procesar solo ese camión y obtener su resumen rápido

    # Llamar a la función principal
    run_quality_analysis(base_camiones_path, output_reports_dir, best_trucks_file, camiones_a_procesar_list=CAMIONES_PARA_PRUEBA)