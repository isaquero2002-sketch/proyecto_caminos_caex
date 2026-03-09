import csv
import glob
import os
from openpyxl import load_workbook


OUTPUT_FORMAT = "csv"  
COLUMNAS_CLAVE = ["Latitud", "Longitud", "Altitude (m)", "Fuel rate (L/h)"]

def stream_csv_rows(file_path):
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        # Detecta el separador (coma o punto y coma)
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(f.read(2048))
            f.seek(0)
        except csv.Error:
            # Si falla, asume coma por defecto
            dialect = 'excel'
            f.seek(0)
        
        reader = csv.reader(f, dialect)
        yield from reader

def stream_excel_rows(file_path):
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows():
            yield [cell.value for cell in row]
    finally:
        if 'wb' in locals():
            wb.close()

def _normalize_numeric_string(s):
    """
    Limpia una cadena para que pueda ser convertida a número.
    Ej: "1.914,2500" -> "1914.2500"
    """
    if not isinstance(s, str):
        return s
    s = s.strip()
    # Lógica  para manejar separadores de miles (.) y decimales (,)
    if ',' in s and '.' in s:
        # Caso "1.234,56": el punto es de miles, la coma es decimal.
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        # Caso "1234,56": la coma es decimal.
        s = s.replace(',', '.')
    # Caso "1.234.567" o "-331.560.414...": los puntos son de miles.
    elif s.count('.') > 1:
        s = s.replace('.', '')
    return s

def clean_and_append_stream(file_path, writer, columns_to_clean, write_header=False):
    """
    Lee un archivo fila por fila, lo limpia y escribe las filas válidas en un 'writer' de CSV ya abierto.
    Retorna el conteo de filas antes y después de la limpieza para este archivo.
    """
    print(f"Procesando: {os.path.basename(file_path)}")
    ext = os.path.splitext(file_path)[1].lower()

    # Seleccionar el iterador de filas según la extensión
    try:
        if ext == ".csv":
            row_iterator = stream_csv_rows(file_path)
        elif ext in [".xlsx", ".xls"]:
            row_iterator = stream_excel_rows(file_path)
        else:
            print(f"  -> Formato no soportado: {ext}")
            return 0, 0
    except Exception as e:
        print(f"  -> Error al abrir el archivo: {e}")
        return 0, 0

    try:
        header = next(row_iterator)
        if not header:
            print("  -> Archivo vacío o sin encabezado.")
            return 0, 0
    except StopIteration:
        print("  -> Archivo vacío.")
        return 0, 0

    # Encontrar los índices de las columnas a limpiar
    try:
        indices_a_limpiar = [header.index(col) for col in columns_to_clean if col in header]
    except ValueError as e:
        print(f"  -> Error: La columna '{e.args[0].split()[0]}' no se encuentra en el archivo.")
        return 0, 0

    total_before = 0
    total_after = 0

    if write_header:
        writer.writerow(header) # Escribir el encabezado

    # Obtener los índices de todas las columnas para aplicar la normalización numérica
    all_indices = list(range(len(header)))

    for row in row_iterator:
        total_before += 1
        
        # Asegurarse de que la fila tenga la misma longitud que el encabezado
        if len(row) != len(header):
            continue # Omitir filas malformadas
        row = [_normalize_numeric_string(cell) for cell in row]

        es_valida = True
        for idx in indices_a_limpiar:
            valor = row[idx]
            # La fila no es válida si el valor es nulo, una cadena vacía, o numéricamente cero
            if valor is None or str(valor).strip() == '':
                es_valida = False
                break
            try:
                if float(valor) == 0:
                    es_valida = False
                    break
            except (ValueError, TypeError):
                # Si no se puede convertir a float, no es cero, así que es válido en este contexto
                pass
        
        if es_valida:
            writer.writerow(row)
            total_after += 1

    # Imprimir estadísticas solo para este archivo
    if total_before > 0:
        pct_eliminadas = (total_before - total_after) / total_before * 100
        print(f"  -> Filas procesadas: {total_before:,}, Filas válidas añadidas: {total_after:,} ({pct_eliminadas:.2f}% eliminadas)")
    
    return total_before, total_after

def process_all_trucks(base_path, output_base_path, trucks_to_process, columns_to_clean):
    if not os.path.isdir(base_path):
        print(f" Ruta base no existe: {base_path}")
        return

    if not trucks_to_process: # Si la lista está vacía, procesar todos
        trucks_to_process = [d for d in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, d))]
        print(f"Se procesarán todos los camiones: {len(trucks_to_process)}")

    os.makedirs(output_base_path, exist_ok=True)

    for camion in trucks_to_process:
        camion_path = os.path.join(base_path, camion)
        if not os.path.isdir(camion_path):
            print(f"No existe carpeta del camión: {camion}")
            continue

        output_camion_dir = os.path.join(output_base_path, camion)
        os.makedirs(output_camion_dir, exist_ok=True)

        print(f"\n--- Camión: {camion} ---")
        
        # Buscar archivos CSV y Excel recursivamente
        files_to_process = []
        for ext in ['csv', 'xlsx', 'xls']:
            files_to_process.extend(glob.glob(os.path.join(camion_path, '**', f'*.{ext}'), recursive=True))
        
        if not files_to_process:
            print("  -> No hay archivos .csv/.xlsx")
            continue
        
        #consolidación
        # 1. Definir un único archivo de salida para el camión
        consolidated_output_file = os.path.join(output_camion_dir, f"{camion}_consolidado_limpio.{OUTPUT_FORMAT}")
        if os.path.exists(consolidated_output_file):
            os.remove(consolidated_output_file)

        total_filas_camion_antes = 0
        total_filas_camion_despues = 0
        
        # 2. Abrir el archivo de salida una sola vez
        with open(consolidated_output_file, 'w', newline='', encoding='utf-8') as f_out:
            # MODIFICACIÓN: Forzar comillas en campos no numéricos para mayor robustez
            writer = csv.writer(f_out, quoting=csv.QUOTE_NONNUMERIC)
            
            # 3. Procesar cada archivo y añadirlo al archivo consolidado
            for i, file_path in enumerate(files_to_process):
                # El encabezado solo se escribe para el primer archivo (i == 0)
                filas_antes, filas_despues = clean_and_append_stream(file_path, writer, columns_to_clean, write_header=(i == 0))
                total_filas_camion_antes += filas_antes
                total_filas_camion_despues += filas_despues

        # 4. Imprimir resumen final para el camión
        if total_filas_camion_despues > 0:
            pct = (total_filas_camion_antes - total_filas_camion_despues) / total_filas_camion_antes * 100 if total_filas_camion_antes > 0 else 0
            print(f"\nResumen para {camion}:")
            print(f"  -> Total Filas antes: {total_filas_camion_antes:,} | Total después: {total_filas_camion_despues:,} | Total eliminadas: {total_filas_camion_antes - total_filas_camion_despues:,} ({pct:.2f}%)")
            print(f"  -> Archivo consolidado guardado en: {consolidated_output_file}")
        else:
            print(f"\n  -> No quedaron filas válidas para el camión {camion}; no se generó archivo de salida.")
            if os.path.exists(consolidated_output_file):
                os.remove(consolidated_output_file)

if __name__ == "__main__":
    # Rutas
    project_base_path = r"C:\Users\icquerov\OneDrive - Anglo American\Desktop\Proyecto_Caminos"
    base_trucks_path = os.path.join(project_base_path, "camiones")
    output_cleaned_path = os.path.join(project_base_path, "outputs", "datos_limpios")

    # Camión específico o todos
    input_camion = input("Ingresa el nombre del camión a procesar (o deja en blanco para procesar todos): ").strip()
    CAMIONES = [input_camion] if input_camion else []

    # Columnas que deben ser no nulas y != 0 (ajusta si no tienes Altitud)
    COLUMNAS_CLAVE = ["Latitud", "Longitud", "Altitude (m)", "Fuel rate (L/h)"]

    process_all_trucks(base_trucks_path, output_cleaned_path, CAMIONES, COLUMNAS_CLAVE)

    print("Proceso completado.")